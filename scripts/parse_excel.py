"""
Parse CommonIngredients.xlsx and DishDescriptions.xlsx into structured JSON.

CommonIngredients.xlsx:
  - Sheet1: 377 rows mapping ingredient ID -> name + market commonality
  - Columns: id, name, CG/CD/CE/CS (common markets), ES/ED/EE/EG (exotic markets)

DishDescriptions.xlsx:
  - DishDescriptions sheet: dish_name, dish_pk, mltext_pk, type, descriptions in en/da/de/lv/et/lt/es
  - DishNotes sheet: same structure but with notes/guides per dish
  - CookingComponents sheet: cooking module templates (en + shortcut + da)
  - RecipeNotes sheet: composed full recipe instructions

Outputs excel_parsed.json with all parsed data and summary statistics.

Usage:
    python3 scripts/parse_excel.py
"""

import json
import os
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.join(SCRIPT_DIR, "..")
INGREDIENTS_PATH = os.path.join(PROJECT_DIR, "CommonIngredients.xlsx")
DISHES_PATH = os.path.join(PROJECT_DIR, "DishDescriptions.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "excel_parsed.json")

# Market column mapping for CommonIngredients
COMMON_COLUMNS = {
    "CG": {"col": 3, "label": "Common_German", "market": "de"},
    "CD": {"col": 4, "label": "Common_Danish", "market": "da"},
    "CE": {"col": 5, "label": "Common_English", "market": "en"},
    "CS": {"col": 6, "label": "Common_Spanish", "market": "es"},
}
EXOTIC_COLUMNS = {
    "ES": {"col": 7, "label": "Exotic_Spanish", "market": "es"},
    "ED": {"col": 8, "label": "Exotic_Danish", "market": "da"},
    "EE": {"col": 9, "label": "Exotic_English", "market": "en"},
    "EG": {"col": 10, "label": "Exotic_German", "market": "de"},
}


def is_formula(value):
    """Check if a cell value is an unresolved Excel formula."""
    return isinstance(value, str) and value.startswith("=")


def safe_str(value):
    """Convert a cell value to string, handling None and formulas."""
    if value is None:
        return None
    if is_formula(value):
        return None  # formula not resolved with data_only
    return str(value).strip() if value else None


def safe_int(value):
    """Convert a cell value to int, handling None, formulas, and floats."""
    if value is None:
        return None
    if is_formula(value):
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def parse_common_ingredients():
    """Parse CommonIngredients.xlsx into structured ingredient data."""
    wb = openpyxl.load_workbook(INGREDIENTS_PATH, data_only=True)
    ws = wb["Sheet1"]

    ingredients = {}
    skipped = 0
    formula_cells = 0
    markets = {"de", "da", "en", "es"}
    common_count = {m: 0 for m in markets}
    exotic_count = {m: 0 for m in markets}
    neither_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        cell_values = [c.value for c in row]

        ingredient_id = safe_int(cell_values[0])
        ingredient_name = safe_str(cell_values[1])

        if ingredient_id is None or ingredient_name is None:
            skipped += 1
            continue

        # Determine which markets this ingredient is common/exotic in
        common_in = []
        exotic_in = []

        for code, info in COMMON_COLUMNS.items():
            val = cell_values[info["col"] - 1]  # 0-indexed
            if is_formula(val):
                formula_cells += 1
            if val and not is_formula(val):
                common_in.append(info["market"])
                common_count[info["market"]] += 1

        for code, info in EXOTIC_COLUMNS.items():
            val = cell_values[info["col"] - 1]  # 0-indexed
            if is_formula(val):
                formula_cells += 1
            if val and not is_formula(val):
                exotic_in.append(info["market"])
                exotic_count[info["market"]] += 1

        if not common_in and not exotic_in:
            neither_count += 1

        ingredients[str(ingredient_id)] = {
            "id": ingredient_id,
            "name": ingredient_name,
            "common_in": sorted(common_in),
            "exotic_in": sorted(exotic_in),
        }

    # Check columns K and L for extra formula cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        for cell in row[10:12]:  # columns K-L (0-indexed 10-11)
            if cell.value and is_formula(str(cell.value)):
                formula_cells += 1

    wb.close()

    summary = {
        "total_ingredients": len(ingredients),
        "skipped_rows": skipped,
        "formula_cells_encountered": formula_cells,
        "common_per_market": common_count,
        "exotic_per_market": exotic_count,
        "neither_common_nor_exotic": neither_count,
    }

    return ingredients, summary


def parse_dish_descriptions():
    """Parse DishDescriptions.xlsx (all sheets) into structured data."""
    # Load with data_only=True to get computed values where possible
    wb = openpyxl.load_workbook(DISHES_PATH, data_only=True)
    # Also load without data_only to detect formulas
    wb_formulas = openpyxl.load_workbook(DISHES_PATH, data_only=False)

    result = {
        "dish_descriptions": [],
        "dish_notes": [],
        "cooking_components": [],
        "recipe_notes": [],
    }
    stats = {
        "formula_cells": 0,
        "sheets_parsed": [],
    }

    # --- DishDescriptions sheet ---
    ws = wb["DishDescriptions"]
    ws_f = wb_formulas["DishDescriptions"]
    lang_cols = {"en": 5, "da": 6, "de": 7, "lv": 8, "et": 9, "lt": 10, "es": 11}
    # 0-indexed: en=5, da=6, de=7, lv=8, et=9, lt=10, es=11

    seen_dish_pks = set()
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2
    ):
        vals = [c.value for c in row]
        f_row = list(ws_f.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        f_vals = [c.value for c in f_row]

        dish_name = safe_str(vals[0])
        dish_pk = safe_int(vals[1])
        mltext_pk = safe_int(vals[2])

        # Column D (Type) is often a formula
        dish_type = safe_str(vals[3])
        if dish_type is None and f_vals[3] and is_formula(str(f_vals[3])):
            stats["formula_cells"] += 1
            dish_type = "__formula__"

        # Column E (New text) is often a formula
        new_text = safe_str(vals[4])
        if new_text is None and f_vals[4] and is_formula(str(f_vals[4])):
            stats["formula_cells"] += 1

        if dish_name is None and dish_pk is None:
            continue

        descriptions = {}
        for lang, col_idx in lang_cols.items():
            text = safe_str(vals[col_idx])
            if text:
                descriptions[lang] = text

        entry = {
            "dish_name": dish_name,
            "dish_pk": dish_pk,
            "mltext_pk": mltext_pk,
            "type": dish_type if dish_type != "__formula__" else None,
            "descriptions": descriptions,
        }
        result["dish_descriptions"].append(entry)
        if dish_pk is not None:
            seen_dish_pks.add(dish_pk)

    stats["sheets_parsed"].append({
        "name": "DishDescriptions",
        "rows_parsed": len(result["dish_descriptions"]),
        "unique_dish_pks": len(seen_dish_pks),
        "languages_available": list(lang_cols.keys()),
    })

    # --- DishNotes sheet ---
    ws = wb["DishNotes"]
    ws_f = wb_formulas["DishNotes"]
    note_lang_cols = {"en": 8, "da": 9, "de": 10, "lv": 11, "et": 12}
    # Also columns for proposals: 6=New proposal (en), 7=New proposal (da)

    seen_note_pks = set()
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2
    ):
        vals = [c.value for c in row]
        f_row = list(ws_f.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        f_vals = [c.value for c in f_row]

        dish_name = safe_str(vals[0])
        dish_pk = safe_int(vals[1])
        mltext_pk = safe_int(vals[2])

        # Type column (D) - often a formula
        dish_type = safe_str(vals[3])
        if dish_type is None and f_vals[3] and is_formula(str(f_vals[3])):
            stats["formula_cells"] += 1

        if dish_name is None and dish_pk is None:
            continue

        # Group column
        group = safe_str(vals[5])

        # New proposal columns
        new_proposal_en = safe_str(vals[6])
        new_proposal_da = safe_str(vals[7])

        notes = {}
        for lang, col_idx in note_lang_cols.items():
            text = safe_str(vals[col_idx])
            if text:
                notes[lang] = text

        entry = {
            "dish_name": dish_name,
            "dish_pk": dish_pk,
            "mltext_pk": mltext_pk,
            "type": dish_type,
            "group": group,
            "new_proposal_en": new_proposal_en,
            "new_proposal_da": new_proposal_da,
            "notes": notes,
        }
        result["dish_notes"].append(entry)
        if dish_pk is not None:
            seen_note_pks.add(dish_pk)

    stats["sheets_parsed"].append({
        "name": "DishNotes",
        "rows_parsed": len(result["dish_notes"]),
        "unique_dish_pks": len(seen_note_pks),
        "languages_available": list(note_lang_cols.keys()),
    })

    # --- CookingComponents sheet ---
    ws = wb["CookingComponents"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        module_name = safe_str(row[0])
        full_text_en = safe_str(row[1])
        shortcut_en = safe_str(row[2])
        shortcut_da = safe_str(row[4]) if len(row) > 4 else None

        if module_name is None and full_text_en is None and shortcut_en is None:
            continue

        entry = {
            "module": module_name,
            "full_text_en": full_text_en,
            "shortcut_en": shortcut_en,
            "shortcut_da": shortcut_da,
        }
        result["cooking_components"].append(entry)

    stats["sheets_parsed"].append({
        "name": "CookingComponents",
        "rows_parsed": len(result["cooking_components"]),
    })

    # --- RecipeNotes sheet ---
    ws = wb["RecipeNotes"]
    ws_f = wb_formulas["RecipeNotes"]
    recipe_formula_count = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2
    ):
        vals = [c.value for c in row]
        f_row = list(ws_f.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        f_vals = [c.value for c in f_row]

        recipe_name = safe_str(vals[0])

        # Column B is often a formula (concatenating cooking components)
        full_recipe_en = safe_str(vals[1])
        is_b_formula = f_vals[1] and is_formula(str(f_vals[1]))
        if is_b_formula:
            recipe_formula_count += 1

        shortcut_en = safe_str(vals[2])
        recipe_id = safe_int(vals[3])
        shortcut_da = safe_str(vals[4])

        if recipe_name is None:
            continue

        entry = {
            "recipe_name": recipe_name,
            "full_recipe_en": full_recipe_en,
            "full_recipe_is_formula": is_b_formula,
            "shortcut_en": shortcut_en,
            "recipe_id": recipe_id,
            "shortcut_da": shortcut_da,
        }
        result["recipe_notes"].append(entry)

    stats["formula_cells"] += recipe_formula_count
    stats["sheets_parsed"].append({
        "name": "RecipeNotes",
        "rows_parsed": len(result["recipe_notes"]),
        "formula_composed_recipes": recipe_formula_count,
    })

    wb.close()
    wb_formulas.close()

    return result, stats


def main():
    print("=" * 60)
    print("Gastrowheel Excel Parser")
    print("=" * 60)

    # --- Parse CommonIngredients ---
    print("\n[1/2] Parsing CommonIngredients.xlsx...")
    ingredients, ing_summary = parse_common_ingredients()

    print(f"  Total ingredients: {ing_summary['total_ingredients']}")
    print(f"  Skipped rows: {ing_summary['skipped_rows']}")
    print(f"  Formula cells encountered: {ing_summary['formula_cells_encountered']}")
    print(f"  Neither common nor exotic: {ing_summary['neither_common_nor_exotic']}")
    print(f"  Common per market: {ing_summary['common_per_market']}")
    print(f"  Exotic per market: {ing_summary['exotic_per_market']}")

    # --- Parse DishDescriptions ---
    print("\n[2/2] Parsing DishDescriptions.xlsx...")
    dishes, dish_stats = parse_dish_descriptions()

    for sheet_info in dish_stats["sheets_parsed"]:
        print(f"  {sheet_info['name']}: {sheet_info['rows_parsed']} rows")
        if "unique_dish_pks" in sheet_info:
            print(f"    Unique dish PKs: {sheet_info['unique_dish_pks']}")
        if "languages_available" in sheet_info:
            print(f"    Languages: {', '.join(sheet_info['languages_available'])}")
        if "formula_composed_recipes" in sheet_info:
            print(f"    Formula-composed recipes: {sheet_info['formula_composed_recipes']}")
    print(f"  Total formula cells across all sheets: {dish_stats['formula_cells']}")

    # --- Build output ---
    output = {
        "common_ingredients": {
            "summary": ing_summary,
            "ingredients": ingredients,
        },
        "dish_descriptions": {
            "summary": dish_stats,
            "descriptions": dishes["dish_descriptions"],
            "notes": dishes["dish_notes"],
            "cooking_components": dishes["cooking_components"],
            "recipe_notes": dishes["recipe_notes"],
        },
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nJSON written to: {OUTPUT_PATH}")

    # --- Sample entries ---
    print("\n" + "=" * 60)
    print("SAMPLE ENTRIES")
    print("=" * 60)

    print("\n--- CommonIngredients (first 5) ---")
    for ing_id, ing in list(ingredients.items())[:5]:
        common = ", ".join(ing["common_in"]) if ing["common_in"] else "none"
        exotic = ", ".join(ing["exotic_in"]) if ing["exotic_in"] else "none"
        print(f"  [{ing['id']}] {ing['name']}")
        print(f"       Common in: {common} | Exotic in: {exotic}")

    print("\n--- DishDescriptions (first 3) ---")
    for entry in dishes["dish_descriptions"][:3]:
        langs = list(entry["descriptions"].keys())
        print(f"  [{entry['dish_pk']}] {entry['dish_name']} (type: {entry['type']})")
        print(f"       Languages: {', '.join(langs)}")
        # Show English description snippet
        en_desc = entry["descriptions"].get("en", "")
        if en_desc:
            snippet = en_desc[:80] + "..." if len(en_desc) > 80 else en_desc
            print(f"       en: {snippet}")

    print("\n--- CookingComponents (first 3) ---")
    for entry in dishes["cooking_components"][:3]:
        print(f"  Module: {entry['module']}")
        if entry["full_text_en"]:
            snippet = entry["full_text_en"][:80]
            print(f"       Full: {snippet}...")
        if entry["shortcut_en"]:
            print(f"       Shortcut: {entry['shortcut_en'][:80]}")

    print("\n--- RecipeNotes (first 3) ---")
    for entry in dishes["recipe_notes"][:3]:
        formula_note = " [composed from formulas]" if entry["full_recipe_is_formula"] else ""
        print(f"  {entry['recipe_name']} (id: {entry['recipe_id']}){formula_note}")
        if entry["full_recipe_en"]:
            snippet = entry["full_recipe_en"][:100] + "..." if len(entry["full_recipe_en"]) > 100 else entry["full_recipe_en"]
            print(f"       Full: {snippet}")
        if entry["shortcut_en"]:
            print(f"       Shortcut: {entry['shortcut_en'][:80]}")


if __name__ == "__main__":
    main()
